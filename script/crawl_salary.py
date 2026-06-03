import argparse
import dataclasses
import datetime as dt
import difflib
import hashlib
import json
import os
import random
import re
import statistics
import time
import urllib.parse
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import pymysql
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright
import yaml


@dataclass(frozen=True)
class SalaryInfo:
    min_month: Optional[float]
    max_month: Optional[float]
    avg_month: Optional[float]
    pay_months: Optional[int]
    annualized_avg: Optional[float]
    unit: Optional[str]


@dataclass(frozen=True)
class JobRecord:
    platform: str
    city: str
    keyword: str
    page_no: int
    job_title: Optional[str]
    company_name: Optional[str]
    salary_text: Optional[str]
    salary_min_month: Optional[float]
    salary_max_month: Optional[float]
    salary_avg_month: Optional[float]
    pay_months: Optional[int]
    salary_annualized_avg: Optional[float]
    location_text: Optional[str]
    education: Optional[str]
    experience_text: Optional[str]
    experience_year_min: Optional[float]
    experience_year_max: Optional[float]
    skills_text: Optional[str]
    job_url: Optional[str]
    crawled_at: dt.datetime
    raw: Dict[str, Any]

    def job_hash(self) -> str:
        """生成用于去重的稳定哈希，优先使用链接，否则使用关键字段拼接。"""
        base = (self.job_url or "").strip()
        if base:
            key = f"{self.platform}|{base}"
        else:
            key = "|".join(
                [
                    self.platform,
                    self.city,
                    self.keyword,
                    (self.job_title or "").strip(),
                    (self.company_name or "").strip(),
                    (self.salary_text or "").strip(),
                ]
            )
        return hashlib.sha1(key.encode("utf-8")).hexdigest()


def parse_salary_text(text: Optional[str]) -> SalaryInfo:
    """解析常见薪资字符串，统一输出为“人民币/月”的区间与平均值，并计算年化平均。"""
    if not text:
        return SalaryInfo(None, None, None, None, None, None)

    s = str(text).strip()
    if not s:
        return SalaryInfo(None, None, None, None, None, None)

    s = s.replace(" ", "")
    pay_months = 12

    m_pay = re.search(r"·(\d{2})薪", s)
    if m_pay:
        try:
            pay_months = int(m_pay.group(1))
        except ValueError:
            pay_months = 12

    s_main = s.split("·", 1)[0]

    if any(x in s_main for x in ["面议", "暂无", "N/A"]):
        return SalaryInfo(None, None, None, pay_months, None, None)

    def to_float(num: str) -> Optional[float]:
        try:
            return float(num)
        except ValueError:
            return None

    def normalize_range(a: Optional[float], b: Optional[float]) -> Tuple[Optional[float], Optional[float]]:
        if a is None and b is None:
            return None, None
        if a is None:
            return b, b
        if b is None:
            return a, a
        return (a, b) if a <= b else (b, a)

    unit = None
    min_month = None
    max_month = None

    m_k = re.match(r"(?P<a>\d+(\.\d+)?)-(?P<b>\d+(\.\d+)?)(?P<u>[Kk])$", s_main)
    if m_k:
        unit = "K/月"
        a = to_float(m_k.group("a"))
        b = to_float(m_k.group("b"))
        if a is not None:
            a *= 1000.0
        if b is not None:
            b *= 1000.0
        min_month, max_month = normalize_range(a, b)

    if min_month is None and max_month is None:
        m_wan = re.match(r"(?P<a>\d+(\.\d+)?)-(?P<b>\d+(\.\d+)?)(?P<u>万)$", s_main)
        if m_wan:
            unit = "万/月"
            a = to_float(m_wan.group("a"))
            b = to_float(m_wan.group("b"))
            if a is not None:
                a *= 10000.0
            if b is not None:
                b *= 10000.0
            min_month, max_month = normalize_range(a, b)

    if min_month is None and max_month is None:
        m_day = re.match(r"(?P<a>\d+(\.\d+)?)-(?P<b>\d+(\.\d+)?)(?P<u>元/天)$", s_main)
        if m_day:
            unit = "元/天"
            a = to_float(m_day.group("a"))
            b = to_float(m_day.group("b"))
            if a is not None:
                a *= 21.75
            if b is not None:
                b *= 21.75
            min_month, max_month = normalize_range(a, b)

    if min_month is None and max_month is None:
        m_year = re.match(r"(?P<a>\d+(\.\d+)?)-(?P<b>\d+(\.\d+)?)(?P<u>万/年)$", s_main)
        if m_year:
            unit = "万/年"
            a = to_float(m_year.group("a"))
            b = to_float(m_year.group("b"))
            if a is not None:
                a = a * 10000.0 / 12.0
            if b is not None:
                b = b * 10000.0 / 12.0
            min_month, max_month = normalize_range(a, b)

    if min_month is None and max_month is None:
        m_rmb_range = re.match(r"(?P<a>\d+(\.\d+)?)-(?P<b>\d+(\.\d+)?)(?P<u>元(/月)?)$", s_main)
        if m_rmb_range:
            unit = "元/月"
            a = to_float(m_rmb_range.group("a"))
            b = to_float(m_rmb_range.group("b"))
            min_month, max_month = normalize_range(a, b)

    if min_month is None and max_month is None:
        m_single_k = re.match(r"(?P<a>\d+(\.\d+)?)(?P<u>[Kk])$", s_main)
        if m_single_k:
            unit = "K/月"
            a = to_float(m_single_k.group("a"))
            if a is not None:
                a *= 1000.0
            min_month, max_month = normalize_range(a, a)

    if min_month is None and max_month is None:
        m_single_w = re.match(r"(?P<a>\d+(\.\d+)?)(?P<u>万)$", s_main)
        if m_single_w:
            unit = "万/月"
            a = to_float(m_single_w.group("a"))
            if a is not None:
                a *= 10000.0
            min_month, max_month = normalize_range(a, a)

    if min_month is None and max_month is None:
        m_single_rmb = re.match(r"(?P<a>\d+(\.\d+)?)(?P<u>元(/月)?)$", s_main)
        if m_single_rmb:
            unit = "元/月"
            a = to_float(m_single_rmb.group("a"))
            min_month, max_month = normalize_range(a, a)

    avg_month = None
    if min_month is not None and max_month is not None:
        avg_month = (min_month + max_month) / 2.0

    annualized_avg = None
    if avg_month is not None and pay_months:
        annualized_avg = avg_month * float(pay_months)

    return SalaryInfo(min_month, max_month, avg_month, pay_months, annualized_avg, unit)

def parse_experience_years(text: Optional[str]) -> Tuple[Optional[float], Optional[float]]:
    """从经验要求文本中提取年限区间。"""
    if not text:
        return None, None
    s = str(text).strip()
    if not s:
        return None, None
    if any(k in s for k in ["经验不限", "不限", "无经验"]):
        return None, None

    s = s.replace("年以上", "+年").replace("年及以上", "+年")
    m_range = re.search(r"(\d+)\s*-\s*(\d+)\s*年", s)
    if m_range:
        return float(m_range.group(1)), float(m_range.group(2))
    m_plus = re.search(r"(\d+)\s*\\+\\s*年", s)
    if m_plus:
        v = float(m_plus.group(1))
        return v, None
    m_single = re.search(r"(\\d+)\\s*年", s)
    if m_single:
        v = float(m_single.group(1))
        return v, v
    return None, None


def normalize_education(text: Optional[str]) -> Optional[str]:
    """规范化学历字段，无法识别时返回原文或None。"""
    if not text:
        return None
    s = str(text).strip()
    if not s:
        return None
    for k in ["博士", "硕士", "本科", "大专", "中专", "高中", "初中"]:
        if k in s:
            return k
    if "不限" in s:
        return "不限"
    return s

def expand_keywords(keywords: Sequence[str]) -> List[str]:
    """将配置中的岗位关键词做规则展开（例如初/中/高级、经理/主管/专员），并去重保持顺序。"""
    expanded: List[str] = []
    seen = set()
    for kw in keywords:
        for item in expand_one_keyword(str(kw).strip()):
            s = item.strip()
            if not s or s in seen:
                continue
            seen.add(s)
            expanded.append(s)
    return expanded


def expand_one_keyword(keyword: str) -> List[str]:
    """展开单个岗位关键词，支持括号级别与&/、分隔的后缀角色组合。"""
    s = str(keyword).strip()
    if not s:
        return []

    s = s.replace("（", "(").replace("）", ")")
    s = s.replace("，", ",").replace("、", ",").replace("/", ",")

    role_suffixes = ["专员", "主管", "经理", "总监"]
    level_prefixes = ["高级", "中级", "初级", "资深"]

    m = re.search(r"\(([^)]{1,50})\)", s)
    if m:
        inside = m.group(1)
        parts = [p.strip() for p in inside.split(",") if p.strip()]
        base = (s[: m.start()] + s[m.end() :]).strip()
        base = base.strip(" -_")
        out: List[str] = []
        for p in parts:
            if p in level_prefixes:
                out.append(f"{p}{base}")
            else:
                out.append(f"{base}{p}")
        return out or [s]

    tokens = [t.strip() for t in re.split(r"[&,+]", s) if t.strip()]
    if len(tokens) <= 1:
        return [s]

    def is_suffix_only(t: str) -> bool:
        return t in role_suffixes

    def strip_suffix(text: str) -> str:
        for suf in role_suffixes:
            if text.endswith(suf):
                return text[: -len(suf)]
        return text

    out: List[str] = []
    last_full: Optional[str] = None
    for t in tokens:
        if is_suffix_only(t) and last_full:
            base = strip_suffix(last_full)
            out.append(base + t)
        else:
            out.append(t)
            last_full = t
    return out or [s]


def classify_role(text: Optional[str]) -> Tuple[Optional[str], Optional[str]]:
    """从岗位标题中识别统一岗位与级别（初/中/高/资深/专员/主管/经理/总监等）。"""
    if not text:
        return None, None
    s = str(text).strip()
    if not s:
        return None, None

    s = s.replace("（", "(").replace("）", ")")
    s = re.sub(r"\s+", "", s)

    level_prefixes = ["资深", "高级", "中级", "初级"]
    role_suffixes = ["总监", "经理", "主管", "专员", "助理"]

    level = None
    for p in level_prefixes:
        if p in s:
            level = p
            s = s.replace(p, "")
            break

    suffix_level = None
    for suf in role_suffixes:
        if s.endswith(suf):
            suffix_level = suf
            s = s[: -len(suf)]
            break

    s = re.sub(r"\([^)]*\)", "", s).strip("-_ ")
    base = s or None
    final_level = suffix_level or level
    return base, final_level


def filter_records_by_keyword(records: Sequence[JobRecord], keyword: str, city: Optional[str] = None) -> List[JobRecord]:
    """按关键词推断约束对结果做轻量过滤，降低模糊搜索带来的噪音，必要时自动降级以避免全空。"""
    kept, _stats = filter_records_with_reason_stats(records, keyword=keyword, city=city)
    return kept


def filter_records_with_reason_stats(
    records: Sequence[JobRecord],
    keyword: str,
    city: Optional[str] = None,
    match_terms: Optional[Sequence[str]] = None,
    match_threshold: float = 0.4,
    relax: bool = False,
) -> Tuple[List[JobRecord], Dict[str, Any]]:
    """对岗位记录做关键词过滤，并输出用于定位 kept=0 的原因统计。"""
    base, level = classify_role(keyword)
    city_raw = str(city).strip() if city else ""
    city_text = city_raw
    if city_text and re.fullmatch(r"\d{2,12}", city_text):
        city_text = ""

    base = base or ""
    core_base = base
    for suf in ["工程师", "专员", "主管", "经理", "总监", "助理", "顾问"]:
        if core_base.endswith(suf):
            core_base = core_base[: -len(suf)]
            break
    core_base = core_base.strip()

    def norm(s: str) -> str:
        s = str(s or "").strip()
        s = s.replace("（", "(").replace("）", ")")
        s = re.sub(r"\s+", "", s)
        return s

    core_base_n = norm(core_base)
    level_n = norm(level or "")
    match_terms_n = []
    if match_terms:
        for t in match_terms:
            tn = norm(str(t))
            if tn and tn not in match_terms_n:
                match_terms_n.append(tn)
    if core_base_n and core_base_n not in match_terms_n:
        match_terms_n.insert(0, core_base_n)

    strict_out: List[JobRecord] = []
    loose_out: List[JobRecord] = []
    dropped_empty_title = 0
    dropped_city = 0
    dropped_base = 0
    dropped_level = 0
    samples: List[str] = []

    for r in records:
        title_raw = (r.job_title or "").strip()
        if not title_raw:
            dropped_empty_title += 1
            continue
        title = norm(title_raw)
        if city_text:
            loc = (r.location_text or "").strip()
            if loc and city_text not in loc:
                dropped_city += 1
                if len(samples) < 5:
                    samples.append(f"title={title_raw} loc={loc}")
                continue
        ok_base = True
        if match_terms_n:
            ok_base = False
            for term in match_terms_n:
                if len(term) < 2:
                    continue
                if term in title:
                    ok_base = True
                    break
                if match_threshold and difflib.SequenceMatcher(None, term, title).ratio() >= float(match_threshold):
                    ok_base = True
                    break
        if not ok_base:
            dropped_base += 1
            if len(samples) < 5:
                samples.append(f"title={title_raw} loc={(r.location_text or '').strip()}")
            continue

        if relax:
            loose_out.append(r)
        else:
            loose_out.append(r)
            if level_n and level_n in title:
                strict_out.append(r)
            elif not level_n:
                strict_out.append(r)
            else:
                dropped_level += 1

    kept = loose_out if relax else (strict_out if strict_out else loose_out)
    used_mode = "relax" if relax else ("strict" if strict_out else "loose")
    stats = {
        "extracted": len(records),
        "kept": len(kept),
        "kept_strict": len(strict_out),
        "kept_loose": len(loose_out),
        "used_mode": used_mode,
        "city_raw": city_raw,
        "city_filter_applied": bool(city_text),
        "core_base": core_base_n,
        "level": level_n,
        "match_terms": match_terms_n,
        "match_threshold": match_threshold,
        "dropped_empty_title": dropped_empty_title,
        "dropped_city": dropped_city,
        "dropped_base": dropped_base,
        "dropped_level": dropped_level,
        "sample_dropped": samples,
    }
    return kept, stats


def build_stealth_context_kwargs() -> Dict[str, Any]:
    """生成较温和的浏览器上下文参数，降低被识别为自动化的概率。"""
    viewport_candidates = [(1920, 1080), (1366, 768), (1440, 900), (1536, 864)]
    w, h = random.choice(viewport_candidates)
    ua = (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/125.0.0.0 Safari/537.36"
    )
    return {
        "locale": "zh-CN",
        "timezone_id": "Asia/Shanghai",
        "viewport": {"width": w, "height": h},
        "user_agent": ua,
        "bypass_csp": True,
    }


def add_stealth_init_script(context) -> None:
    """向页面注入初始化脚本，尽量减少自动化特征。"""
    context.add_init_script(
        """
        Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
        Object.defineProperty(navigator, 'languages', {get: () => ['zh-CN', 'zh']});
        Object.defineProperty(navigator, 'plugins', {get: () => [1,2,3,4,5]});
        window.chrome = window.chrome || { runtime: {} };
        """
    )


def normalize_url(base: str, href: Optional[str]) -> Optional[str]:
    """将相对链接或协议相对链接规范化为绝对URL。"""
    if not href:
        return None
    u = href.strip()
    if not u:
        return None
    if u.startswith("//"):
        return "https:" + u
    if u.startswith("http://") or u.startswith("https://"):
        return u
    return urllib.parse.urljoin(base, u)


def safe_text(el) -> Optional[str]:
    """从元素中提取文本，失败时返回None。"""
    try:
        t = el.inner_text()
        t = t.strip()
        return t or None
    except Exception:
        return None


def query_first_text(root, selectors: Sequence[str]) -> Optional[str]:
    """按候选选择器顺序取第一个命中的文本。"""
    for sel in selectors:
        try:
            el = root.query_selector(sel)
            if el:
                t = safe_text(el)
                if t:
                    return t
        except Exception:
            continue
    return None


def detect_challenge(page) -> Optional[str]:
    """检测常见安全校验/验证码页面特征，命中时返回原因。"""
    try:
        title = (page.title() or "").strip()
    except Exception:
        title = ""
    try:
        content = page.content()
    except Exception:
        content = ""
    try:
        url = (page.url or "").strip()
    except Exception:
        url = ""

    if url and any(x in url for x in ["/web/user", "passport-zp", "from=passport", "fromUrl="]):
        return "需要登录"

    if title and any(x in title for x in ["注册登录", "登录", "请先登录"]):
        return "需要登录"

    if url == "about:blank":
        try:
            if "<body></body>" in content or content.strip() == "<html><head></head><body></body></html>":
                return "页面空白"
        except Exception:
            return "页面空白"

    for kw in ["安全验证", "滑动验证", "验证码", "人机识别", "访问受限", "验证中心"]:
        if kw in title or kw in content:
            return kw
    return None


def boss_normalize_city(city_code_or_name: str) -> str:
    """将常见城市中文名映射为BOSS直聘city参数，未命中时原样返回。"""
    s = str(city_code_or_name).strip()
    if not s:
        return s
    if re.fullmatch(r"\d{6,12}", s):
        return s
    mapping = {
        "北京": "101010100",
        "上海": "101020100",
        "广州": "101280100",
        "深圳": "101280600",
        "杭州": "101210100",
        "南京": "101190100",
        "苏州": "101190400",
        "成都": "101270100",
        "重庆": "101040100",
        "武汉": "101200100",
        "西安": "101110100",
        "天津": "101030100",
        "郑州": "101180100",
        "长沙": "101250100",
        "合肥": "101220100",
        "厦门": "101230200",
        "福州": "101230100",
        "济南": "101120100",
        "青岛": "101120200",
        "沈阳": "101070100",
        "大连": "101070200",
        "昆明": "101290100",
        "南宁": "101300100",
        "南昌": "101240100",
        "石家庄": "101090100",
        "太原": "101100100",
        "无锡": "101190200",
        "宁波": "101210400",
        "东莞": "101281600",
        "佛山": "101280800",
    }
    return mapping.get(s, s)


_ZL_CITY_CODE_CACHE: Optional[Dict[str, str]] = None


def zhilian_normalize_city(city_code_or_name: str) -> str:
    """将城市中文名映射为智联城市code（如703），未命中时原样返回。"""
    s = str(city_code_or_name).strip()
    if not s:
        return s
    if re.fullmatch(r"\\d{2,6}", s):
        return s
    global _ZL_CITY_CODE_CACHE
    if _ZL_CITY_CODE_CACHE is None:
        _ZL_CITY_CODE_CACHE = load_zhilian_city_code_map()
    return _ZL_CITY_CODE_CACHE.get(s, s)


def boss_build_joblist_api_url(keyword: str, city_code_or_name: str, page_no: int) -> str:
    """构造BOSS直聘岗位列表接口URL（同源 fetch 用）。"""
    params = {
        "scene": "1",
        "query": keyword,
        "city": boss_normalize_city(city_code_or_name),
        "page": str(page_no),
        "pageSize": "30",
    }
    return "https://www.zhipin.com/wapi/zpgeek/search/joblist.json?" + urllib.parse.urlencode(params)


def boss_cookie_header(cookies: Sequence[Dict[str, Any]]) -> str:
    """将Playwright cookies结构转换为HTTP Cookie请求头字符串。"""
    parts: List[str] = []
    for c in cookies:
        try:
            name = str(c.get("name") or "").strip()
            value = str(c.get("value") or "").strip()
        except Exception:
            continue
        if name and value:
            parts.append(f"{name}={value}")
    return "; ".join(parts)


def boss_fetch_joblist_by_cookies(api_url: str, cookies: Sequence[Dict[str, Any]], user_agent: str) -> Optional[Dict[str, Any]]:
    """使用浏览器cookie在HTTP层请求BOSS岗位列表接口，避免影响已打开的浏览器页面。"""
    import urllib.request

    cookie_str = boss_cookie_header(cookies)
    if not cookie_str:
        return None
    headers = {
        "User-Agent": user_agent,
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "zh-CN,zh;q=0.9",
        "Referer": "https://www.zhipin.com/web/geek/jobs",
        "Origin": "https://www.zhipin.com",
        "X-Requested-With": "XMLHttpRequest",
        "Cookie": cookie_str,
    }
    req = urllib.request.Request(api_url, headers=headers, method="GET")
    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            raw = resp.read()
        return json.loads(raw.decode("utf-8", errors="ignore"))
    except Exception:
        return None


def boss_payload_error(payload: Dict[str, Any]) -> Optional[str]:
    """解析BOSS接口返回中的错误信息，正常时返回None。"""
    try:
        code = payload.get("code")
        if code is None:
            return None
        code_str = str(code).strip()
        if code_str in {"0", ""}:
            return None
        msg = payload.get("message") or payload.get("msg") or payload.get("error") or ""
        msg = str(msg).strip()
        return f"code={code_str} msg={msg}" if msg else f"code={code_str}"
    except Exception:
        return None


def boss_payload_code(payload: Dict[str, Any]) -> Optional[str]:
    """提取BOSS接口返回码，缺失时返回None。"""
    try:
        code = payload.get("code")
        if code is None:
            return None
        return str(code).strip()
    except Exception:
        return None


def sleep_between(min_s: float, max_s: float, verbose: bool, label: str) -> None:
    """按区间随机睡眠，用于控制请求节奏。"""
    try:
        a = float(min_s)
        b = float(max_s)
    except Exception:
        return
    if a <= 0 and b <= 0:
        return
    if b < a:
        a, b = b, a
    s = random.uniform(a, b) if b > 0 else 0.0
    if s <= 0:
        return
    if verbose:
        print(f"[sleep] {label} seconds={round(s, 2)}")
    time.sleep(s)


def boss_is_rate_limited(payload: Dict[str, Any]) -> bool:
    """判断BOSS接口是否处于访问频率限制/账号限制状态。"""
    try:
        msg = payload.get("message") or payload.get("msg") or payload.get("error") or ""
    except Exception:
        msg = ""
    s = str(msg).strip()
    if not s:
        return False
    for k in ["访问过于频繁", "请求过于频繁", "操作过于频繁", "被限制", "限制访问", "稍后再试"]:
        if k in s:
            return True
    return False


def boss_is_env_abnormal(payload: Dict[str, Any]) -> bool:
    """判断BOSS接口是否返回“环境异常/风控”类错误码。"""
    code = boss_payload_code(payload) or ""
    return code in {"37", "36", "38"}


def boss_get_user_agent_from_context(context) -> str:
    """尽量从当前CDP上下文获取真实User-Agent，用于HTTP层请求。"""
    try:
        for p0 in reversed(getattr(context, "pages", [])):
            try:
                if p0 and (not p0.is_closed()) and (p0.url or "").strip().startswith("http"):
                    ua = p0.evaluate("() => navigator.userAgent")
                    if isinstance(ua, str) and ua.strip():
                        return ua.strip()
            except Exception:
                continue
    except Exception:
        pass
    return str(build_stealth_context_kwargs().get("user_agent") or "Mozilla/5.0").strip()


def boss_close_tagged_pages(context, keep_page=None) -> int:
    """关闭脚本为BOSS自动化流程创建的历史页面，避免页面堆积占用内存。"""
    closed = 0
    try:
        pages = list(getattr(context, "pages", []))
    except Exception:
        pages = []
    for p0 in pages:
        if keep_page is not None and p0 is keep_page:
            continue
        try:
            if getattr(p0, "_jd_scrapy_tag", "") != "boss_auto":
                continue
        except Exception:
            continue
        try:
            if not p0.is_closed():
                p0.close()
                closed += 1
        except Exception:
            continue
    return closed


def boss_mark_page(page) -> None:
    """为脚本创建的BOSS页面打标，便于后续清理。"""
    try:
        setattr(page, "_jd_scrapy_tag", "boss_auto")
    except Exception:
        pass


def boss_recover_login_state(context, verbose: bool, page=None) -> None:
    """当BOSS接口触发风控时，尝试在CDP浏览器内恢复登录态/完成验证。"""
    created = False
    if page is None:
        boss_close_tagged_pages(context, keep_page=None)
        created = True
        try:
            page = context.new_page()
            boss_mark_page(page)
        except Exception:
            page = None
    try:
        if page is None:
            return
        page.goto("https://www.zhipin.com/web/geek/jobs", wait_until="domcontentloaded", timeout=60000)
        for _ in range(120):
            reason = detect_challenge(page)
            if not reason:
                return
            if verbose:
                print(f"[boss] challenge={reason}，请在已打开的Chrome中完成验证/刷新页面，脚本将自动继续...")
            if reason == "需要登录":
                try:
                    page.locator("a:has-text('登录')").first.click(timeout=1500)
                except Exception:
                    try:
                        page.locator("button:has-text('登录')").first.click(timeout=1500)
                    except Exception:
                        pass
            time.sleep(2)
            try:
                page.reload(wait_until="domcontentloaded", timeout=60000)
            except Exception:
                pass
    finally:
        if created and page:
            try:
                page.close()
            except Exception:
                pass


def boss_prepare_api_page(context, verbose: bool):
    """为BOSS接口请求准备一个稳定页面，用于同源fetch以降低风控概率。"""
    page = None
    try:
        for p0 in reversed(getattr(context, "pages", [])):
            try:
                u = (p0.url or "").strip()
                if (not p0.is_closed()) and u and ("zhipin.com" in u) and (u != "about:blank") and (not u.startswith("chrome://")):
                    page = p0
                    break
            except Exception:
                continue
    except Exception:
        page = None
    if page is None:
        boss_close_tagged_pages(context, keep_page=None)
        page = context.new_page()
        boss_mark_page(page)
    try:
        page.goto("https://www.zhipin.com/web/geek/jobs", wait_until="domcontentloaded", timeout=60000)
    except Exception:
        pass
    for _ in range(180):
        try:
            u = (page.url or "").strip()
            if u and ("zhipin.com" not in u) and (not u.startswith("about:")) and (not u.startswith("chrome://")):
                page.goto("https://www.zhipin.com/web/geek/jobs", wait_until="domcontentloaded", timeout=60000)
        except Exception:
            pass
        reason = detect_challenge(page)
        if not reason:
            return page
        if verbose:
            print(f"[boss] challenge={reason}，请在已打开的Chrome中完成验证/刷新页面，脚本将自动继续...")
        time.sleep(2)
        try:
            page.reload(wait_until="domcontentloaded", timeout=60000)
        except Exception:
            pass
    return page


def boss_fetch_joblist_by_page_fetch(page, api_url: str) -> Optional[Dict[str, Any]]:
    """在浏览器页面内用fetch请求BOSS接口，尽量复用浏览器网络栈与登录态。"""
    try:
        result = page.evaluate(
            """
            async (u) => {
              const resp = await fetch(u, {
                method: 'GET',
                credentials: 'include',
                headers: {
                  'accept': 'application/json, text/plain, */*',
                  'x-requested-with': 'XMLHttpRequest'
                }
              });
              const text = await resp.text();
              return { status: resp.status, text };
            }
            """,
            api_url,
        )
        if not isinstance(result, dict):
            return None
        text = result.get("text")
        if not isinstance(text, str) or not text.strip():
            return None
        return json.loads(text)
    except Exception:
        return None


def load_zhilian_city_code_map() -> Dict[str, str]:
    """从智联基础数据接口加载城市code映射表。"""
    import urllib.request

    url = "https://fe-api.zhaopin.com/c/i/search/base/data"
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0", "Referer": "https://sou.zhaopin.com/"})
    with urllib.request.urlopen(req, timeout=30) as resp:
        data = json.loads(resp.read())
    root = data.get("data", {}) if isinstance(data, dict) else {}
    out: Dict[str, str] = {}
    for key in ["hotCity", "allCity"]:
        items = root.get(key)
        if not isinstance(items, list):
            continue
        for item in items:
            if not isinstance(item, dict):
                continue
            code = item.get("code")
            name = item.get("name")
            if code and name:
                out[str(name)] = str(code)
            sub = item.get("sublist")
            if isinstance(sub, list):
                for sub_item in sub:
                    if not isinstance(sub_item, dict):
                        continue
                    code2 = sub_item.get("code")
                    name2 = sub_item.get("name")
                    if code2 and name2:
                        out[str(name2)] = str(code2)
    return out


class BossCrawler:
    def __init__(
        self,
        headless: bool,
        proxy: Optional[str],
        slow_mo_ms: int,
        min_sleep: float,
        max_sleep: float,
        verbose: bool,
        user_data_dir: Optional[str],
        channel: Optional[str],
        cdp_endpoint: Optional[str],
        random_pause: bool,
        match_terms: Optional[Sequence[str]] = None,
        match_threshold: float = 0.4,
        match_relax: bool = False,
    ):
        self.headless = headless
        self.proxy = proxy
        self.slow_mo_ms = slow_mo_ms
        self.min_sleep = min_sleep
        self.max_sleep = max_sleep
        self.verbose = verbose
        self.user_data_dir = user_data_dir
        self.channel = channel
        self.cdp_endpoint = cdp_endpoint
        self.random_pause = random_pause
        self.match_terms = list(match_terms) if match_terms else None
        self.match_threshold = float(match_threshold)
        self.match_relax = bool(match_relax)

    def build_url(self, keyword: str, city_code_or_name: str, page_no: int) -> str:
        """构造BOSS直聘搜索URL，city建议传城市code（如101010100）。"""
        q = urllib.parse.quote(keyword)
        city = urllib.parse.quote(boss_normalize_city(city_code_or_name))
        return f"https://www.zhipin.com/web/geek/jobs?query={q}&city={city}&page={page_no}"

    def extract_jobs(self, page, city: str, keyword: str, page_no: int) -> List[JobRecord]:
        """从BOSS结果页抽取岗位记录，尽量做选择器兜底。"""
        cards = []
        for sel in [
            "div.job-card-wrapper",
            "li.job-card-wrapper",
            "div.job-card",
            "div.search-job-result div.job-card-wrapper",
        ]:
            try:
                cards = page.query_selector_all(sel)
                if cards:
                    break
            except Exception:
                continue

        records: List[JobRecord] = []
        for card in cards:
            title = query_first_text(card, [".job-name", ".job-title", "a .job-name"])
            salary = query_first_text(card, [".salary", ".job-salary", "span.salary"])
            company = query_first_text(card, [".company-name", ".job-card-company-name", ".company-text"])
            href = None
            try:
                a = card.query_selector("a")
                if a:
                    href = a.get_attribute("href")
            except Exception:
                href = None
            url = normalize_url("https://www.zhipin.com", href)

            salary_info = parse_salary_text(salary)
            raw = {
                "query_keyword": keyword,
                "title": title,
                "salary": salary,
                "company": company,
            }
            records.append(
                JobRecord(
                    platform="boss",
                    city=city,
                    keyword=keyword,
                    page_no=page_no,
                    job_title=title,
                    company_name=company,
                    salary_text=salary,
                    salary_min_month=salary_info.min_month,
                    salary_max_month=salary_info.max_month,
                    salary_avg_month=salary_info.avg_month,
                    pay_months=salary_info.pay_months,
                    salary_annualized_avg=salary_info.annualized_avg,
                    location_text=None,
                    education=None,
                    experience_text=None,
                    experience_year_min=None,
                    experience_year_max=None,
                    skills_text=None,
                    job_url=url,
                    crawled_at=dt.datetime.now(),
                    raw=raw,
                )
            )
        return records

    def extract_jobs_from_joblist_json(self, payload: Dict[str, Any], city: str, keyword: str, page_no: int) -> List[JobRecord]:
        """从BOSS接口 joblist.json 的返回中抽取岗位记录。"""
        root: Any = payload
        if isinstance(root, dict):
            root = root.get("zpData") or root.get("data") or root

        job_list: Any = None
        if isinstance(root, dict):
            job_list = root.get("jobList") or root.get("job_list") or root.get("list")
        if not isinstance(job_list, list):
            job_list = []

        records: List[JobRecord] = []
        for item in job_list:
            if not isinstance(item, dict):
                continue
            title = item.get("jobName") or item.get("jobTitle") or item.get("name")
            salary = item.get("salaryDesc") or item.get("salary") or item.get("salaryText")
            company = item.get("brandName") or item.get("companyName") or item.get("brand")
            education = normalize_education(item.get("jobDegree") or item.get("degreeName") or item.get("degree"))
            experience_text = item.get("jobExperience") or item.get("experienceName") or item.get("experience")
            experience_year_min, experience_year_max = parse_experience_years(experience_text)

            labels = item.get("jobLabels") or item.get("skills") or item.get("labels")
            skills: List[str] = []
            if isinstance(labels, list):
                for x in labels:
                    if isinstance(x, str) and x.strip():
                        skills.append(x.strip())
            skills_text = ";".join(dict.fromkeys(skills)) if skills else None

            url = item.get("jobUrl") or item.get("detailUrl") or item.get("jobDetailUrl")
            if not url:
                enc_job_id = item.get("encryptJobId") or item.get("jobId")
                if enc_job_id:
                    url = f"https://www.zhipin.com/job_detail/{enc_job_id}.html"

            salary_info = parse_salary_text(salary)
            raw = {"query_keyword": keyword, "joblist_item": item}
            records.append(
                JobRecord(
                    platform="boss",
                    city=city,
                    keyword=keyword,
                    page_no=page_no,
                    job_title=title,
                    company_name=company,
                    salary_text=salary,
                    salary_min_month=salary_info.min_month,
                    salary_max_month=salary_info.max_month,
                    salary_avg_month=salary_info.avg_month,
                    pay_months=salary_info.pay_months,
                    salary_annualized_avg=salary_info.annualized_avg,
                    location_text=item.get("jobArea") or item.get("locationName") or item.get("area"),
                    education=education,
                    experience_text=experience_text,
                    experience_year_min=experience_year_min,
                    experience_year_max=experience_year_max,
                    skills_text=skills_text,
                    job_url=url,
                    crawled_at=dt.datetime.now(),
                    raw=raw,
                )
            )
        return records

    def crawl_cdp_batch(
        self,
        cities: Sequence[str],
        keyword_items: Sequence[Dict[str, Any]],
        pages_per_city: int,
    ) -> Tuple[List[JobRecord], List[Dict[str, Any]]]:
        """在CDP模式下复用同一个浏览器连接批量抓取多个关键词，避免频繁重连导致跳回新建标签页。"""
        if not self.cdp_endpoint:
            raise RuntimeError("crawl_cdp_batch 仅支持 cdp_endpoint 模式")
        results: List[JobRecord] = []
        missing: List[Dict[str, Any]] = []
        with sync_playwright() as p:
            if self.verbose:
                print(f"[boss] connect_cdp={self.cdp_endpoint}")
            browser = p.chromium.connect_over_cdp(self.cdp_endpoint)
            context = browser.contexts[0] if browser.contexts else browser.new_context(**build_stealth_context_kwargs())

            ua = boss_get_user_agent_from_context(context)
            if self.min_sleep < 6.0:
                self.min_sleep = 6.0
            if self.max_sleep < 12.0:
                self.max_sleep = 12.0

            api_page = boss_prepare_api_page(context, verbose=self.verbose)
            for item in keyword_items:
                keyword = str(item.get("keyword") or "").strip()
                search_keyword = str(item.get("search_keyword") or "").strip() or keyword
                match_terms = item.get("match_terms") if isinstance(item.get("match_terms"), list) else None

                if not keyword:
                    continue

                for city in cities:
                    extracted_any = False
                    kept_any = False
                    last_error: Optional[str] = None
                    for page_no in range(1, pages_per_city + 1):
                        api_url = boss_build_joblist_api_url(search_keyword, city, page_no)
                        if self.verbose:
                            print(f"[boss] joblist_http url={api_url}")
                        payload: Optional[Dict[str, Any]] = None
                        last_err: Optional[str] = None
                        for attempt in range(1, 4):
                            if api_page is None or api_page.is_closed():
                                api_page = boss_prepare_api_page(context, verbose=self.verbose)
                            payload = boss_fetch_joblist_by_page_fetch(api_page, api_url)
                            if not payload:
                                cookies = context.cookies("https://www.zhipin.com")
                                payload = boss_fetch_joblist_by_cookies(api_url, cookies=cookies, user_agent=ua)
                            if not payload:
                                last_err = "CDP接口请求失败：cookie为空或HTTP请求失败"
                            else:
                                err = boss_payload_error(payload)
                                if not err:
                                    last_err = None
                                    break
                                last_err = f"BOSS接口返回异常：{err}"
                                if boss_is_env_abnormal(payload):
                                    boss_recover_login_state(context, verbose=self.verbose, page=api_page)
                                time.sleep(random.uniform(3.0, 8.0))
                                continue
                            if attempt < 3:
                                boss_recover_login_state(context, verbose=self.verbose, page=api_page)
                                time.sleep(random.uniform(3.0, 8.0))

                        if not payload:
                            last_error = "CDP接口请求失败：请确保Chrome已登录BOSS且可正常浏览 zhipin.com 页面"
                            break
                        if last_err:
                            last_error = f"{last_err}（请确认用于CDP的Chrome已登录）"
                            break

                        page_records = self.extract_jobs_from_joblist_json(payload, city=city, keyword=keyword, page_no=page_no)
                        extracted_any = extracted_any or (len(page_records) > 0)
                        kept, stats = filter_records_with_reason_stats(
                            page_records,
                            keyword=keyword,
                            city=city,
                            match_terms=match_terms,
                            match_threshold=self.match_threshold,
                            relax=self.match_relax,
                        )
                        kept_any = kept_any or (len(kept) > 0)
                        if self.verbose:
                            print(f"[boss] joblist_api=Y extracted={len(page_records)} kept={len(kept)}")
                            if len(page_records) > 0 and len(kept) == 0:
                                print(f"[filter] platform=boss kw={keyword} stats={json.dumps(stats, ensure_ascii=False)}")
                        results.extend(kept)
                        time.sleep(random.uniform(self.min_sleep, self.max_sleep))

                    if last_error:
                        missing.append({"platform": "boss", "city": city, "keyword": keyword, "reason": f"异常: {last_error}"})
                    elif not extracted_any:
                        missing.append({"platform": "boss", "city": city, "keyword": keyword, "reason": "无结果"})
                    elif not kept_any:
                        missing.append({"platform": "boss", "city": city, "keyword": keyword, "reason": "过滤后为0"})

            return results, missing

    def crawl_cdp_one(
        self,
        context,
        api_page,
        ua: str,
        cities: Sequence[str],
        keyword: str,
        search_keyword: str,
        match_terms: Optional[Sequence[str]],
        pages_per_city: int,
        boss_gap_min: float,
        boss_gap_max: float,
    ) -> Tuple[List[JobRecord], List[Dict[str, Any]]]:
        """在已建立CDP连接的前提下抓取单个关键词，并在每次BOSS请求后做间隔。"""
        results: List[JobRecord] = []
        missing: List[Dict[str, Any]] = []
        kw = str(keyword).strip()
        q_kw = str(search_keyword).strip() or kw
        if not kw:
            return results, missing

        for city in cities:
            extracted_any = False
            kept_any = False
            last_error: Optional[str] = None
            for page_no in range(1, pages_per_city + 1):
                api_url = boss_build_joblist_api_url(q_kw, city, page_no)
                if self.verbose:
                    print(f"[boss] joblist_http url={api_url}")
                payload: Optional[Dict[str, Any]] = None
                last_err: Optional[str] = None
                for attempt in range(1, 4):
                    if api_page is None or api_page.is_closed():
                        api_page = boss_prepare_api_page(context, verbose=self.verbose)
                    payload = boss_fetch_joblist_by_page_fetch(api_page, api_url)
                    if not payload:
                        cookies = context.cookies("https://www.zhipin.com")
                        payload = boss_fetch_joblist_by_cookies(api_url, cookies=cookies, user_agent=ua)
                    if not payload:
                        last_err = "CDP接口请求失败：cookie为空或HTTP请求失败"
                    else:
                        if boss_is_rate_limited(payload):
                            last_err = "BOSS触发请求频率限制/账号限制"
                            sleep_between(60.0, 180.0, verbose=self.verbose, label="boss_backoff")
                            break
                        err = boss_payload_error(payload)
                        if not err:
                            last_err = None
                            break
                        last_err = f"BOSS接口返回异常：{err}"
                        if boss_is_env_abnormal(payload):
                            boss_recover_login_state(context, verbose=self.verbose, page=api_page)
                        sleep_between(3.0, 8.0, verbose=self.verbose, label="boss_retry")
                        continue
                    if attempt < 3:
                        boss_recover_login_state(context, verbose=self.verbose, page=api_page)
                        sleep_between(3.0, 8.0, verbose=self.verbose, label="boss_retry")

                if not payload:
                    last_error = "CDP接口请求失败：请确保Chrome已登录BOSS且可正常浏览 zhipin.com 页面"
                    break
                if last_err:
                    last_error = f"{last_err}（请确认用于CDP的Chrome已登录）"
                    break

                page_records = self.extract_jobs_from_joblist_json(payload, city=city, keyword=kw, page_no=page_no)
                extracted_any = extracted_any or (len(page_records) > 0)
                kept, stats = filter_records_with_reason_stats(
                    page_records,
                    keyword=kw,
                    city=city,
                    match_terms=match_terms,
                    match_threshold=self.match_threshold,
                    relax=self.match_relax,
                )
                kept_any = kept_any or (len(kept) > 0)
                if self.verbose:
                    print(f"[boss] joblist_api=Y extracted={len(page_records)} kept={len(kept)}")
                    if len(page_records) > 0 and len(kept) == 0:
                        print(f"[filter] platform=boss kw={kw} stats={json.dumps(stats, ensure_ascii=False)}")
                results.extend(kept)
                sleep_between(boss_gap_min, boss_gap_max, verbose=self.verbose, label="boss_gap")

            if last_error:
                missing.append({"platform": "boss", "city": city, "keyword": kw, "reason": f"异常: {last_error}"})
            elif not extracted_any:
                missing.append({"platform": "boss", "city": city, "keyword": kw, "reason": "无结果"})
            elif not kept_any:
                missing.append({"platform": "boss", "city": city, "keyword": kw, "reason": "过滤后为0"})

        return results, missing

    def crawl(self, cities: Sequence[str], keyword: str, pages_per_city: int, search_keyword: Optional[str] = None) -> List[JobRecord]:
        """执行BOSS直聘采集。"""
        results: List[JobRecord] = []
        q_kw = str(search_keyword).strip() if search_keyword else keyword
        with sync_playwright() as p:
            launch_kwargs: Dict[str, Any] = {"headless": self.headless, "slow_mo": self.slow_mo_ms}
            if self.proxy:
                launch_kwargs["proxy"] = {"server": self.proxy}
            if self.channel:
                launch_kwargs["channel"] = self.channel
            launch_kwargs["args"] = [
                "--disable-blink-features=AutomationControlled",
                "--disable-features=IsolateOrigins,site-per-process",
            ]
            launch_kwargs["ignore_default_args"] = ["--enable-automation"]

            context = None
            browser = None
            should_close = True
            if self.cdp_endpoint:
                if self.verbose:
                    print(f"[boss] connect_cdp={self.cdp_endpoint}")
                browser = p.chromium.connect_over_cdp(self.cdp_endpoint)
                should_close = False
                context = browser.contexts[0] if browser.contexts else browser.new_context(**build_stealth_context_kwargs())
            elif self.user_data_dir:
                context = p.chromium.launch_persistent_context(
                    self.user_data_dir, **launch_kwargs, **build_stealth_context_kwargs()
                )
                add_stealth_init_script(context)
            else:
                browser = p.chromium.launch(**launch_kwargs)
                context = browser.new_context(**build_stealth_context_kwargs())
                add_stealth_init_script(context)

            if self.cdp_endpoint:
                ua = boss_get_user_agent_from_context(context)
                if self.min_sleep < 6.0:
                    self.min_sleep = 6.0
                if self.max_sleep < 12.0:
                    self.max_sleep = 12.0
                api_page = boss_prepare_api_page(context, verbose=self.verbose)
                for city in cities:
                    for page_no in range(1, pages_per_city + 1):
                        api_url = boss_build_joblist_api_url(q_kw, city, page_no)
                        if self.verbose:
                            print(f"[boss] joblist_http url={api_url}")
                        payload: Optional[Dict[str, Any]] = None
                        last_err: Optional[str] = None
                        for attempt in range(1, 4):
                            if api_page is None or api_page.is_closed():
                                api_page = boss_prepare_api_page(context, verbose=self.verbose)
                            payload = boss_fetch_joblist_by_page_fetch(api_page, api_url)
                            if not payload:
                                cookies = context.cookies("https://www.zhipin.com")
                                payload = boss_fetch_joblist_by_cookies(api_url, cookies=cookies, user_agent=ua)
                            if not payload:
                                last_err = "CDP接口请求失败：cookie为空或HTTP请求失败"
                            else:
                                err = boss_payload_error(payload)
                                if not err:
                                    last_err = None
                                    break
                                last_err = f"BOSS接口返回异常：{err}"
                                if boss_is_env_abnormal(payload):
                                    boss_recover_login_state(context, verbose=self.verbose, page=api_page)
                                time.sleep(random.uniform(3.0, 8.0))
                                continue
                            if attempt < 3:
                                boss_recover_login_state(context, verbose=self.verbose, page=api_page)
                                time.sleep(random.uniform(3.0, 8.0))
                        if not payload:
                            raise RuntimeError("CDP接口请求失败：请确保Chrome已登录BOSS且可正常浏览 zhipin.com 页面")
                        if last_err:
                            raise RuntimeError(f"{last_err}（请确认用于CDP的Chrome已登录）")
                        page_records = self.extract_jobs_from_joblist_json(payload, city=city, keyword=keyword, page_no=page_no)
                        kept, stats = filter_records_with_reason_stats(
                            page_records,
                            keyword=keyword,
                            city=city,
                            match_terms=self.match_terms,
                            match_threshold=self.match_threshold,
                            relax=self.match_relax,
                        )
                        if self.verbose:
                            print(f"[boss] joblist_api=Y extracted={len(page_records)} kept={len(kept)}")
                            if len(page_records) > 0 and len(kept) == 0:
                                print(f"[filter] platform=boss kw={keyword} stats={json.dumps(stats, ensure_ascii=False)}")
                        results.extend(kept)
                        time.sleep(random.uniform(self.min_sleep, self.max_sleep))
                if browser and should_close:
                    browser.close()
                return results

            if "page" not in locals() or page is None:
                page = None
            if page is None:
                try:
                    for p0 in reversed(getattr(context, "pages", [])):
                        try:
                            if not p0.is_closed() and p0.url and "zhipin.com" in p0.url:
                                page = p0
                                break
                        except Exception:
                            continue
                except Exception:
                    page = None
            if page is None:
                page = context.new_page()
            if self.verbose:
                try:
                    page.on("pageerror", lambda exc: print(f"[boss] pageerror={exc}"))
                    page.on(
                        "console",
                        lambda msg: print(f"[boss] console[{msg.type}] {msg.text}")
                        if msg.type in {"warning", "error"}
                        else None,
                    )
                except Exception:
                    pass

            def pick_active_page() -> None:
                nonlocal page
                try:
                    if page.is_closed():
                        page = context.pages[-1]
                except Exception:
                    pass
                try:
                    if page.url == "about:blank":
                        for p2 in reversed(context.pages):
                            try:
                                if not p2.is_closed() and p2.url and "zhipin.com" in p2.url and p2.url != "about:blank":
                                    page = p2
                                    break
                            except Exception:
                                continue
                except Exception:
                    pass

            for city in cities:
                for page_no in range(1, pages_per_city + 1):
                    url = self.build_url(q_kw, city, page_no)
                    if self.verbose:
                        print(f"[boss] city={city} kw={keyword} page={page_no}/{pages_per_city} url={url}")
                    api_payload: Optional[Dict[str, Any]] = None
                    resp = None
                    if not api_payload:
                        try:
                            with page.expect_response(
                                lambda r: "/wapi/zpgeek/search/joblist.json" in (r.url or "") and r.status == 200,
                                timeout=25000,
                            ) as resp_info:
                                try:
                                    resp = page.goto(url, wait_until="domcontentloaded", timeout=60000)
                                except Exception as e:
                                    if self.verbose:
                                        print(f"[boss] goto_error={e}")
                                    try:
                                        page.evaluate("u => { window.location.href = u; }", url)
                                    except Exception:
                                        pass
                            try:
                                api_payload = resp_info.value.json()
                            except Exception:
                                api_payload = None
                        except PlaywrightTimeoutError:
                            try:
                                resp = page.goto(url, wait_until="domcontentloaded", timeout=60000)
                            except Exception as e:
                                if self.verbose:
                                    print(f"[boss] goto_timeout_then_error={e}")
                    if self.verbose:
                        try:
                            status = resp.status if resp else None
                            print(f"[boss] resp_status={status}")
                        except Exception:
                            pass
                    pick_active_page()
                    if self.verbose:
                        try:
                            print(f"[boss] after_goto url={page.url} title={page.title()}")
                        except Exception:
                            pass
                    reason = detect_challenge(page)
                    if reason:
                        if self.headless:
                            raise RuntimeError(
                                f"BOSS触发校验：{reason}。建议设置 crawl.user_data_dir 并用 headless=false 先登录/过验证一次"
                            )
                        if self.verbose:
                            print(f"[boss] challenge={reason}，等待你在浏览器中完成登录/验证...")
                        while True:
                            time.sleep(2)
                            pick_active_page()
                            reason2 = detect_challenge(page)
                            if not reason2:
                                break
                    try:
                        page.wait_for_selector("div.job-card-wrapper, li.job-card-wrapper", timeout=15000)
                    except PlaywrightTimeoutError:
                        pass
                    pick_active_page()
                    reason = detect_challenge(page)
                    if reason and self.headless:
                        raise RuntimeError(
                            f"BOSS触发校验：{reason}。建议设置 crawl.user_data_dir 并用 headless=false 先登录/过验证一次"
                        )
                    if reason and not self.headless:
                        if self.verbose:
                            print(f"[boss] challenge={reason}，等待你在浏览器中完成登录/验证...")
                        while True:
                            time.sleep(2)
                            pick_active_page()
                            reason2 = detect_challenge(page)
                            if not reason2:
                                break
                    if self.verbose:
                        try:
                            cookies = context.cookies("https://www.zhipin.com")
                            print(f"[boss] final_url={page.url} cookies={len(cookies)} pages={len(context.pages)}")
                            if page.url == "about:blank":
                                for i, p2 in enumerate(context.pages):
                                    try:
                                        print(f"[boss] page[{i}] url={p2.url} title={p2.title() if not p2.is_closed() else '-'}")
                                    except Exception:
                                        continue
                        except Exception:
                            pass
                    page.wait_for_timeout(random.randint(800, 1600))
                    pick_active_page()
                    if api_payload:
                        page_records = self.extract_jobs_from_joblist_json(api_payload, city=city, keyword=keyword, page_no=page_no)
                        if self.verbose:
                            print(f"[boss] joblist_api=Y")
                    else:
                        page_records = self.extract_jobs(page, city=city, keyword=keyword, page_no=page_no)
                        if self.verbose:
                            print(f"[boss] joblist_api=N")
                    kept, stats = filter_records_with_reason_stats(
                        page_records,
                        keyword=keyword,
                        city=city,
                        match_terms=self.match_terms,
                        match_threshold=self.match_threshold,
                        relax=self.match_relax,
                    )
                    if self.verbose:
                        print(f"[boss] extracted={len(page_records)} kept={len(kept)}")
                        if len(page_records) > 0 and len(kept) == 0:
                            print(f"[filter] platform=boss kw={keyword} stats={json.dumps(stats, ensure_ascii=False)}")
                        if len(page_records) == 0:
                            try:
                                html = page.content()
                                print(f"[boss] html_head={html[:400].replace(chr(10),' ')}")
                            except Exception:
                                pass
                    results.extend(kept)
                    time.sleep(random.uniform(self.min_sleep, self.max_sleep))

            if not self.headless and self.random_pause:
                pause_s = random.randint(1, 60)
                if self.verbose:
                    print(f"[boss] pause_seconds={pause_s}")
                time.sleep(pause_s)

            try:
                if should_close:
                    context.close()
            except Exception:
                pass
            if browser and should_close:
                browser.close()
        return results


class ZhilianCrawler:
    def __init__(
        self,
        headless: bool,
        proxy: Optional[str],
        slow_mo_ms: int,
        min_sleep: float,
        max_sleep: float,
        verbose: bool,
        user_data_dir: Optional[str],
        match_terms: Optional[Sequence[str]] = None,
        match_threshold: float = 0.4,
        match_relax: bool = False,
    ):
        self.headless = headless
        self.proxy = proxy
        self.slow_mo_ms = slow_mo_ms
        self.min_sleep = min_sleep
        self.max_sleep = max_sleep
        self.verbose = verbose
        self.user_data_dir = user_data_dir
        self.match_terms = list(match_terms) if match_terms else None
        self.match_threshold = float(match_threshold)
        self.match_relax = bool(match_relax)

    def build_url(self, keyword: str, city_name: str, page_no: int) -> str:
        """构造智联招聘搜索URL，city建议使用城市code以确保筛选生效。"""
        q = urllib.parse.quote(keyword)
        c = urllib.parse.quote(zhilian_normalize_city(city_name))
        return f"https://sou.zhaopin.com/?jl={c}&kw={q}&p={page_no}"

    def extract_jobs(self, page, city: str, keyword: str, page_no: int) -> List[JobRecord]:
        """从智联结果页抽取岗位记录，尽量做选择器兜底。"""
        cards = []
        for sel in [
            "div.joblist-box__iteminfo",
            "div.joblist-box__item",
            "div.joblist-box__item-content",
            "div.joblist-box__itembox",
            "div.joblist div.item",
        ]:
            try:
                cards = page.query_selector_all(sel)
                if cards:
                    break
            except Exception:
                continue

        records: List[JobRecord] = []
        for card in cards:
            title = query_first_text(card, ["a.jobinfo__name", ".jobinfo__name", "a[href*='jobdetail']"])
            salary = query_first_text(card, [".jobinfo__salary", ".joblist-box__salary", ".salary", ".job-salary"])
            company = query_first_text(card, [".companyinfo__name", ".companyinfo__company", ".joblist-box__company"])
            location_text = query_first_text(
                card,
                [
                    ".jobinfo__other-info-location-image + span",
                    ".jobinfo__other-info-item span",
                ],
            )

            experience_text = None
            education_text = None
            try:
                other_items = card.query_selector_all(".jobinfo__other-info-item")
            except Exception:
                other_items = []
            for el in other_items:
                t = safe_text(el)
                if not t:
                    continue
                if "·" in t:
                    continue
                if any(k in t for k in ["博士", "硕士", "本科", "大专", "中专", "高中", "学历不限", "不限"]):
                    if education_text is None:
                        education_text = t
                    continue
                if "年" in t or "经验" in t or "不限" in t:
                    if experience_text is None:
                        experience_text = t
                    continue

            experience_year_min, experience_year_max = parse_experience_years(experience_text)
            education = normalize_education(education_text)

            skills: List[str] = []
            try:
                for el in card.query_selector_all(".jobinfo__tag .joblist-box__item-tag"):
                    t = safe_text(el)
                    if t:
                        skills.append(t)
            except Exception:
                skills = []
            skills_text = ";".join(dict.fromkeys(skills)) if skills else None

            href = None
            try:
                a = card.query_selector("a.jobinfo__name, a[href*='jobdetail']")
                if a:
                    href = a.get_attribute("href")
            except Exception:
                href = None
            url = normalize_url("https://www.zhaopin.com", href)

            salary_info = parse_salary_text(salary)
            raw = {
                "query_keyword": keyword,
                "title": title,
                "salary": salary,
                "company": company,
                "location": location_text,
                "education": education,
                "experience": experience_text,
                "skills": skills,
            }
            records.append(
                JobRecord(
                    platform="zhilian",
                    city=city,
                    keyword=keyword,
                    page_no=page_no,
                    job_title=title,
                    company_name=company,
                    salary_text=salary,
                    salary_min_month=salary_info.min_month,
                    salary_max_month=salary_info.max_month,
                    salary_avg_month=salary_info.avg_month,
                    pay_months=salary_info.pay_months,
                    salary_annualized_avg=salary_info.annualized_avg,
                    location_text=location_text,
                    education=education,
                    experience_text=experience_text,
                    experience_year_min=experience_year_min,
                    experience_year_max=experience_year_max,
                    skills_text=skills_text,
                    job_url=url,
                    crawled_at=dt.datetime.now(),
                    raw=raw,
                )
            )
        return records

    def crawl_with_page(
        self,
        page,
        cities: Sequence[str],
        keyword: str,
        pages_per_city: int,
        search_keyword: Optional[str] = None,
        zhilian_gap_min: float = 1.0,
        zhilian_gap_max: float = 2.0,
    ) -> List[JobRecord]:
        """执行智联招聘采集。"""
        results: List[JobRecord] = []
        q_kw = str(search_keyword).strip() if search_keyword else keyword
        for city in cities:
            for page_no in range(1, pages_per_city + 1):
                url = self.build_url(q_kw, city, page_no)
                if self.verbose:
                    print(f"[zhilian] city={city} kw={keyword} page={page_no}/{pages_per_city} url={url}")
                page.goto(url, wait_until="domcontentloaded", timeout=60000)
                reason = detect_challenge(page)
                if reason and self.headless:
                    raise RuntimeError(f"智联触发安全校验：{reason}，建议使用 --headless false 运行后手动过验证")
                try:
                    page.wait_for_selector("div.joblist-box__iteminfo, div.jobinfo__top", timeout=15000)
                except PlaywrightTimeoutError:
                    pass
                page.wait_for_timeout(random.randint(800, 1600))
                page_records = self.extract_jobs(page, city=city, keyword=keyword, page_no=page_no)
                kept, stats = filter_records_with_reason_stats(
                    page_records,
                    keyword=keyword,
                    city=city,
                    match_terms=self.match_terms,
                    match_threshold=self.match_threshold,
                    relax=self.match_relax,
                )
                if self.verbose:
                    print(f"[zhilian] extracted={len(page_records)} kept={len(kept)}")
                    if len(page_records) > 0 and len(kept) == 0:
                        print(f"[filter] platform=zhilian kw={keyword} stats={json.dumps(stats, ensure_ascii=False)}")
                results.extend(kept)
                sleep_between(zhilian_gap_min, zhilian_gap_max, verbose=self.verbose, label="zhilian_gap")
        return results

    def crawl(self, cities: Sequence[str], keyword: str, pages_per_city: int, search_keyword: Optional[str] = None) -> List[JobRecord]:
        """执行智联招聘采集。"""
        with sync_playwright() as p:
            launch_kwargs: Dict[str, Any] = {"headless": self.headless, "slow_mo": self.slow_mo_ms}
            if self.proxy:
                launch_kwargs["proxy"] = {"server": self.proxy}

            context = None
            browser = None
            if self.user_data_dir:
                context = p.chromium.launch_persistent_context(self.user_data_dir, **launch_kwargs, **build_stealth_context_kwargs())
                add_stealth_init_script(context)
            else:
                browser = p.chromium.launch(**launch_kwargs)
                context = browser.new_context(**build_stealth_context_kwargs())
                add_stealth_init_script(context)
            page = context.new_page()

            try:
                return self.crawl_with_page(page, cities=cities, keyword=keyword, pages_per_city=pages_per_city, search_keyword=search_keyword)
            finally:
                try:
                    context.close()
                except Exception:
                    pass
                if browser:
                    try:
                        browser.close()
                    except Exception:
                        pass


def ensure_dir(path: str) -> None:
    """确保目录存在，不存在则创建。"""
    os.makedirs(path, exist_ok=True)


def write_excel(records: Sequence[JobRecord], out_path: str, missing: Optional[Sequence[Dict[str, Any]]] = None) -> str:
    """将明细与汇总写入Excel。"""
    wb = Workbook()
    ws_raw = wb.active
    ws_raw.title = "raw"

    raw_headers = [
        "平台",
        "城市",
        "岗位关键词",
        "统一岗位",
        "级别",
        "页码",
        "岗位名称",
        "公司",
        "工作地点",
        "学历",
        "工作经验",
        "经验下限(年)",
        "经验上限(年)",
        "技能要求",
        "薪资文本",
        "月薪下限(元)",
        "月薪上限(元)",
        "月薪均值(元)",
        "薪资月数",
        "年化均值(元)",
        "岗位链接",
        "抓取时间",
    ]
    ws_raw.append(raw_headers)
    for c in range(1, len(raw_headers) + 1):
        ws_raw.cell(row=1, column=c).font = Font(bold=True)

    for r in records:
        base_role, level = classify_role(r.job_title) if r.job_title else (None, None)
        if not base_role:
            base_role, level = classify_role(r.keyword)
        ws_raw.append(
            [
                r.platform,
                r.city,
                r.keyword,
                base_role,
                level,
                r.page_no,
                r.job_title,
                r.company_name,
                r.location_text,
                r.education,
                r.experience_text,
                r.experience_year_min,
                r.experience_year_max,
                r.skills_text,
                r.salary_text,
                r.salary_min_month,
                r.salary_max_month,
                r.salary_avg_month,
                r.pay_months,
                r.salary_annualized_avg,
                r.job_url,
                r.crawled_at.strftime("%Y-%m-%d %H:%M:%S"),
            ]
        )

    for idx, h in enumerate(raw_headers, start=1):
        max_len = len(h)
        for row in ws_raw.iter_rows(min_row=2, min_col=idx, max_col=idx, max_row=min(ws_raw.max_row, 300)):
            v = row[0].value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws_raw.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 60)

    ws_sum = wb.create_sheet("summary")
    sum_headers = ["平台", "城市", "统一岗位", "样本数", "月薪均值(元)", "月薪中位数(元)", "月薪最小(元)", "月薪最大(元)"]
    ws_sum.append(sum_headers)
    for c in range(1, len(sum_headers) + 1):
        ws_sum.cell(row=1, column=c).font = Font(bold=True)

    grouped: Dict[Tuple[str, str, str], List[float]] = {}
    for r in records:
        if r.salary_avg_month is None:
            continue
        base_role, _ = classify_role(r.job_title) if r.job_title else (None, None)
        if not base_role:
            base_role, _ = classify_role(r.keyword)
        if not base_role:
            continue
        key = (r.platform, r.city, base_role)
        grouped.setdefault(key, []).append(float(r.salary_avg_month))

    for (platform, city, base_role), values in sorted(grouped.items()):
        values_sorted = sorted(values)
        ws_sum.append(
            [
                platform,
                city,
                base_role,
                len(values_sorted),
                round(statistics.mean(values_sorted), 2) if values_sorted else None,
                round(statistics.median(values_sorted), 2) if values_sorted else None,
                round(values_sorted[0], 2) if values_sorted else None,
                round(values_sorted[-1], 2) if values_sorted else None,
            ]
        )

    for idx, h in enumerate(sum_headers, start=1):
        ws_sum.column_dimensions[get_column_letter(idx)].width = max(len(h) + 2, 16)

    ws_missing = wb.create_sheet("missing")
    missing_headers = ["平台", "城市", "岗位关键词", "原因"]
    ws_missing.append(missing_headers)
    for c in range(1, len(missing_headers) + 1):
        ws_missing.cell(row=1, column=c).font = Font(bold=True)
    if missing:
        for item in missing:
            ws_missing.append(
                [
                    item.get("platform"),
                    item.get("city"),
                    item.get("keyword"),
                    item.get("reason"),
                ]
            )
    for idx, h in enumerate(missing_headers, start=1):
        ws_missing.column_dimensions[get_column_letter(idx)].width = max(len(h) + 2, 24)

    ensure_dir(os.path.dirname(out_path) or ".")
    try:
        wb.save(out_path)
        return out_path
    except PermissionError:
        base, ext = os.path.splitext(out_path)
        ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        alt = f"{base}_{ts}{ext or '.xlsx'}"
        wb.save(alt)
        return alt


def mysql_connect_from_env() -> Optional[pymysql.Connection]:
    """从环境变量读取MySQL连接信息并建立连接，缺失关键信息则返回None。"""
    host = os.getenv("MYSQL_HOST")
    user = os.getenv("MYSQL_USER")
    password = os.getenv("MYSQL_PASSWORD")
    database = os.getenv("MYSQL_DATABASE")
    port = int(os.getenv("MYSQL_PORT", "3306"))
    if not host or not user or password is None or not database:
        return None
    return pymysql.connect(
        host=host,
        user=user,
        password=password,
        database=database,
        port=port,
        charset="utf8mb4",
        autocommit=True,
    )


def mysql_connect_from_config(mysql_cfg: Dict[str, Any]) -> Optional[pymysql.Connection]:
    """从配置字典读取MySQL连接信息并建立连接，缺失关键信息则返回None。"""
    if not mysql_cfg:
        return None
    enabled = mysql_cfg.get("enabled", False)
    if not enabled:
        return None
    host = mysql_cfg.get("host")
    user = mysql_cfg.get("user")
    password = mysql_cfg.get("password")
    database = mysql_cfg.get("database")
    port = int(mysql_cfg.get("port", 3306))
    if not host or not user or password is None or not database:
        return None
    return pymysql.connect(
        host=str(host),
        user=str(user),
        password=str(password),
        database=str(database),
        port=port,
        charset="utf8mb4",
        autocommit=True,
    )


def load_yaml_config(path: str) -> Dict[str, Any]:
    """加载YAML配置文件，返回字典结构。"""
    with open(path, "r", encoding="utf-8") as f:
        data = yaml.safe_load(f) or {}
    if not isinstance(data, dict):
        raise ValueError("config.yml 顶层必须是字典结构")
    return data


_LLM_KEYWORD_PLAN_CACHE: Dict[str, Dict[str, Any]] = {}


def llm_load_keyword_plan_cache(path: str) -> Dict[str, Dict[str, Any]]:
    """加载岗位关键词规划缓存JSON，返回 keyword -> plan。"""
    p = str(path or "").strip()
    if not p:
        return {}
    if not os.path.exists(p):
        return {}
    try:
        with open(p, "r", encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, dict):
            return {}
        out: Dict[str, Dict[str, Any]] = {}
        for k, v in data.items():
            if not isinstance(k, str) or not k.strip():
                continue
            if not isinstance(v, dict):
                continue
            out[k.strip()] = v
        return out
    except Exception:
        return {}


def llm_save_keyword_plan_cache(path: str, cache: Dict[str, Dict[str, Any]]) -> None:
    """保存岗位关键词规划缓存JSON。"""
    p = str(path or "").strip()
    if not p:
        return
    ensure_dir(os.path.dirname(p) or ".")
    tmp = p + ".tmp"
    payload = json.dumps(cache or {}, ensure_ascii=False, indent=2)
    with open(tmp, "w", encoding="utf-8") as f:
        f.write(payload)
    os.replace(tmp, p)


def llm_keyword_plan_cache_path(llm_cfg: Dict[str, Any]) -> str:
    """获取岗位关键词规划缓存文件路径。"""
    if not isinstance(llm_cfg, dict):
        return os.path.join("conf", "llm_keyword_plan_cache.json")
    p = str(llm_cfg.get("cache_path", "")).strip()
    return p or os.path.join("conf", "llm_keyword_plan_cache.json")


def llm_get_api_key(llm_cfg: Dict[str, Any]) -> Tuple[str, str]:
    """读取大模型API Key（优先环境变量，其次配置直接填key），返回(key, source)。"""
    if not isinstance(llm_cfg, dict) or not llm_cfg:
        return "", "none"
    direct = str(llm_cfg.get("api_key", "")).strip()
    if direct:
        return direct, "config"

    env_name_or_key = str(llm_cfg.get("api_key_env", "DEEPSEEK_API_KEY")).strip() or "DEEPSEEK_API_KEY"
    if env_name_or_key.startswith("sk-") and len(env_name_or_key) >= 20:
        return env_name_or_key, "config"

    v = os.getenv(env_name_or_key, "").strip()
    if v:
        return v, "env"
    return "", "none"


def llm_enabled(llm_cfg: Dict[str, Any]) -> bool:
    """判断大模型是否启用（优先从环境变量读取密钥）。"""
    if not isinstance(llm_cfg, dict) or not llm_cfg:
        return False
    enabled = llm_cfg.get("enabled", False)
    if isinstance(enabled, str):
        enabled = str_to_bool(enabled)
    if not enabled:
        return False
    key, _ = llm_get_api_key(llm_cfg)
    return bool(key)


def llm_deepseek_chat(llm_cfg: Dict[str, Any], messages: List[Dict[str, str]]) -> Optional[str]:
    """调用 DeepSeek OpenAI兼容接口，返回 message.content 文本。"""
    import urllib.request

    api_key, _ = llm_get_api_key(llm_cfg)
    if not api_key:
        return None
    endpoint = str(llm_cfg.get("endpoint", "https://api.deepseek.com/v1/chat/completions")).strip()
    model = str(llm_cfg.get("model", "deepseek-v4")).strip() or "deepseek-v4"
    timeout_s = int(llm_cfg.get("timeout_s", 25))
    temperature = float(llm_cfg.get("temperature", 0.2))

    payload = json.dumps(
        {
            "model": model,
            "messages": messages,
            "temperature": temperature,
        },
        ensure_ascii=False,
    ).encode("utf-8")
    req = urllib.request.Request(
        endpoint,
        data=payload,
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=timeout_s) as resp:
            raw = resp.read().decode("utf-8", errors="ignore")
        data = json.loads(raw) if raw else {}
        choices = data.get("choices") if isinstance(data, dict) else None
        if not isinstance(choices, list) or not choices:
            return None
        msg = choices[0].get("message") if isinstance(choices[0], dict) else None
        content = msg.get("content") if isinstance(msg, dict) else None
        return str(content).strip() if isinstance(content, str) and content.strip() else None
    except Exception:
        return None


def _extract_json_object(text: str) -> Optional[Dict[str, Any]]:
    """从文本中提取第一个JSON对象并解析为字典。"""
    if not text:
        return None
    s = str(text).strip()
    if not s:
        return None
    if s.startswith("{") and s.endswith("}"):
        try:
            obj = json.loads(s)
            return obj if isinstance(obj, dict) else None
        except Exception:
            pass
    m = re.search(r"\{[\s\S]{1,2000}\}", s)
    if not m:
        return None
    try:
        obj = json.loads(m.group(0))
        return obj if isinstance(obj, dict) else None
    except Exception:
        return None


def llm_build_keyword_plan(keyword: str, llm_cfg: Dict[str, Any]) -> Dict[str, Any]:
    """用大模型把岗位关键词转成更通用的招聘搜索词，并给出更宽松的匹配词列表。"""
    kw = str(keyword).strip()
    if not kw:
        return {"search_query": "", "match_terms": []}
    if kw in _LLM_KEYWORD_PLAN_CACHE:
        return _LLM_KEYWORD_PLAN_CACHE[kw]
    if not llm_enabled(llm_cfg):
        return {"search_query": "", "match_terms": []}

    system = (
        "你是招聘网站搜索词优化助手。"
        "给定一个中文岗位关键词，请输出适合在招聘网站通用搜索框使用的更通用搜索词，并给出宽松匹配词列表。"
        "只输出JSON对象，不要输出其他文字。"
    )
    user = (
        "岗位关键词："
        + kw
        + "\n"
        + "请输出JSON，格式："
        + '{"search_query":"...","match_terms":["...","...","..."]}'
        + "\n"
        + "search_query为更通用的搜索词（尽量短，1~6个字）；match_terms为宽松匹配词（3~8个）。"
    )
    content = llm_deepseek_chat(
        llm_cfg,
        messages=[
            {"role": "system", "content": system},
            {"role": "user", "content": user},
        ],
    )
    obj = _extract_json_object(content or "")
    plan = {"search_query": "", "match_terms": []}
    if isinstance(obj, dict):
        sq = obj.get("search_query")
        if isinstance(sq, str) and sq.strip():
            plan["search_query"] = sq.strip()
        mt = obj.get("match_terms")
        if isinstance(mt, list):
            cleaned = []
            for x in mt:
                xs = str(x).strip()
                if xs and xs not in cleaned:
                    cleaned.append(xs)
            plan["match_terms"] = cleaned
    if isinstance(llm_cfg, dict):
        model = str(llm_cfg.get("model", "")).strip()
        if model:
            plan["model"] = model
        plan["updated_at"] = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    _LLM_KEYWORD_PLAN_CACHE[kw] = plan
    return plan



def mysql_ensure_table(conn: pymysql.Connection, table: str) -> None:
    """创建用于落库的明细表（不存在则创建），并补齐新增字段列。"""
    ddl = f"""
    CREATE TABLE IF NOT EXISTS `{table}` (
      `id` BIGINT NOT NULL AUTO_INCREMENT,
      `job_hash` CHAR(40) NOT NULL,
      `platform` VARCHAR(20) NOT NULL,
      `city` VARCHAR(50) NOT NULL,
      `keyword` VARCHAR(100) NOT NULL,
      `page_no` INT NOT NULL,
      `job_title` VARCHAR(200) NULL,
      `company_name` VARCHAR(200) NULL,
      `location_text` VARCHAR(100) NULL,
      `education` VARCHAR(50) NULL,
      `experience_text` VARCHAR(50) NULL,
      `experience_year_min` DECIMAL(6,2) NULL,
      `experience_year_max` DECIMAL(6,2) NULL,
      `skills_text` TEXT NULL,
      `salary_text` VARCHAR(100) NULL,
      `salary_min_month` DECIMAL(12,2) NULL,
      `salary_max_month` DECIMAL(12,2) NULL,
      `salary_avg_month` DECIMAL(12,2) NULL,
      `pay_months` INT NULL,
      `salary_annualized_avg` DECIMAL(14,2) NULL,
      `job_url` VARCHAR(500) NULL,
      `crawled_at` DATETIME NOT NULL,
      `raw_json` LONGTEXT NULL,
      PRIMARY KEY (`id`),
      UNIQUE KEY `uk_job_hash` (`job_hash`),
      KEY `idx_platform_city_kw` (`platform`, `city`, `keyword`)
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
    """
    with conn.cursor() as cur:
        cur.execute(ddl)
    mysql_ensure_columns(conn, table=table)


def mysql_ensure_columns(conn: pymysql.Connection, table: str) -> None:
    """检查并补齐新增字段列，避免历史表结构缺列导致写入失败。"""
    desired = {
        "location_text": "VARCHAR(100) NULL",
        "education": "VARCHAR(50) NULL",
        "experience_text": "VARCHAR(50) NULL",
        "experience_year_min": "DECIMAL(6,2) NULL",
        "experience_year_max": "DECIMAL(6,2) NULL",
        "skills_text": "TEXT NULL",
    }
    sql_cols = """
    SELECT COLUMN_NAME
    FROM information_schema.COLUMNS
    WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = %s;
    """
    with conn.cursor() as cur:
        cur.execute(sql_cols, (table,))
        existing = {row[0] for row in cur.fetchall()}
        for col, typ in desired.items():
            if col in existing:
                continue
            cur.execute(f"ALTER TABLE `{table}` ADD COLUMN `{col}` {typ};")


def mysql_insert_records(conn: pymysql.Connection, table: str, records: Sequence[JobRecord]) -> int:
    """批量写入MySQL，使用job_hash去重。"""
    if not records:
        return 0
    sql = f"""
    INSERT INTO `{table}` (
      job_hash, platform, city, keyword, page_no,
      job_title, company_name,
      location_text, education, experience_text, experience_year_min, experience_year_max, skills_text,
      salary_text,
      salary_min_month, salary_max_month, salary_avg_month, pay_months, salary_annualized_avg,
      job_url, crawled_at, raw_json
    ) VALUES (
      %s, %s, %s, %s, %s,
      %s, %s,
      %s, %s, %s, %s, %s, %s,
      %s,
      %s, %s, %s, %s, %s,
      %s, %s, %s
    )
    ON DUPLICATE KEY UPDATE
      crawled_at = VALUES(crawled_at),
      location_text = VALUES(location_text),
      education = VALUES(education),
      experience_text = VALUES(experience_text),
      experience_year_min = VALUES(experience_year_min),
      experience_year_max = VALUES(experience_year_max),
      skills_text = VALUES(skills_text),
      salary_text = VALUES(salary_text),
      salary_min_month = VALUES(salary_min_month),
      salary_max_month = VALUES(salary_max_month),
      salary_avg_month = VALUES(salary_avg_month),
      pay_months = VALUES(pay_months),
      salary_annualized_avg = VALUES(salary_annualized_avg),
      raw_json = VALUES(raw_json);
    """
    rows = 0
    with conn.cursor() as cur:
        for r in records:
            payload = (
                r.job_hash(),
                r.platform,
                r.city,
                r.keyword,
                r.page_no,
                r.job_title,
                r.company_name,
                r.location_text,
                r.education,
                r.experience_text,
                r.experience_year_min,
                r.experience_year_max,
                r.skills_text,
                r.salary_text,
                r.salary_min_month,
                r.salary_max_month,
                r.salary_avg_month,
                r.pay_months,
                r.salary_annualized_avg,
                r.job_url,
                r.crawled_at.strftime("%Y-%m-%d %H:%M:%S"),
                json.dumps(r.raw, ensure_ascii=False),
            )
            cur.execute(sql, payload)
            rows += 1
    return rows


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    """解析命令行参数。"""
    p = argparse.ArgumentParser()
    p.add_argument("--config", default=None, help="YAML配置文件路径，例如：conf/config.yml")
    p.add_argument("--keyword", required=False, default=None, help="岗位关键词，例如：Python工程师")
    p.add_argument("--cities", required=False, default=None, help="城市列表，逗号分隔，例如：北京,上海,深圳 或 BOSS城市code")
    p.add_argument("--pages", type=int, default=5, help="每个城市抓取页数")
    p.add_argument("--platform", choices=["boss", "zhilian", "both"], default=None, help="抓取平台（不填则使用配置）")
    p.add_argument("--headless", type=str, default=None, help="true/false，遇验证码建议false（不填则使用配置）")
    p.add_argument("--proxy", default=None, help="代理，例如：http://user:pass@host:port")
    p.add_argument("--slowmo", type=int, default=0, help="Playwright slow_mo 毫秒")
    p.add_argument("--sleep-min", type=float, default=2.0, help="翻页间隔最小秒数")
    p.add_argument("--sleep-max", type=float, default=5.0, help="翻页间隔最大秒数")
    p.add_argument("--out", default=None, help="输出Excel路径，默认 output/salary_时间戳.xlsx")
    p.add_argument("--mysql-table", default="job_salary_raw", help="MySQL表名")
    p.add_argument("--save-mysql", type=str, default="false", help="true/false，从环境变量读取MySQL连接并落库")
    p.add_argument("--verbose", type=str, default="false", help="true/false，打印抓取进度日志")
    p.add_argument("--user-data-dir", default=None, help="浏览器用户数据目录，用于保留登录态（BOSS建议配置）")
    p.add_argument("--boss-headless", default=None, help="BOSS单独指定true/false（不填则使用全局headless）")
    p.add_argument("--boss-channel", default=None, help="BOSS浏览器通道，例如 chrome（不填则使用默认）")
    p.add_argument("--boss-cdp", default=None, help="连接已打开的Chrome调试端口，例如 http://127.0.0.1:9222")
    return p.parse_args(argv)


def str_to_bool(v: str) -> bool:
    """将常见字符串转换为布尔值。"""
    return str(v).strip().lower() in {"1", "true", "yes", "y", "on"}


def main(argv: Optional[Sequence[str]] = None) -> int:
    """主入口：按平台抓取数据，导出Excel，可选落库到MySQL。"""
    args = parse_args(argv)
    cfg: Dict[str, Any] = {}
    if args.config:
        cfg = load_yaml_config(str(args.config))

    crawl_cfg = cfg.get("crawl", {}) if isinstance(cfg.get("crawl", {}), dict) else {}
    data_cfg = cfg.get("data", {}) if isinstance(cfg.get("data", {}), dict) else {}
    mysql_cfg = cfg.get("mysql", {}) if isinstance(cfg.get("mysql", {}), dict) else {}
    llm_cfg = cfg.get("llm", {}) if isinstance(cfg.get("llm", {}), dict) else {}
    llm_cache_path = llm_keyword_plan_cache_path(llm_cfg)
    cache_loaded = llm_load_keyword_plan_cache(llm_cache_path)
    if cache_loaded:
        _LLM_KEYWORD_PLAN_CACHE.update(cache_loaded)

    cities_arg = str(args.cities).strip() if args.cities else ""
    cities = [c.strip() for c in cities_arg.split(",") if c.strip()] if cities_arg else []
    if not cities:
        cities = [str(x).strip() for x in data_cfg.get("cities", []) if str(x).strip()]
    if not cities:
        raise SystemExit("cities 不能为空（请通过 --cities 或 config.yml 的 data.cities 提供）")

    keywords: List[str] = []
    if args.keyword:
        k = str(args.keyword).strip()
        if k:
            keywords = [k]
    if not keywords:
        keywords = [str(x).strip() for x in data_cfg.get("keywords", []) if str(x).strip()]

    if not keywords:
        raise SystemExit("keyword 不能为空（请通过 --keyword 或 config.yml 的 data.keywords 提供）")

    keywords = expand_keywords(keywords)

    pages = int(crawl_cfg.get("pages_per_city", args.pages))
    if pages <= 0:
        raise SystemExit("pages 必须大于0")

    max_keywords = crawl_cfg.get("max_keywords", None)
    if max_keywords is not None and str(max_keywords).strip() != "":
        try:
            mk = int(max_keywords)
        except ValueError:
            mk = 0
        if mk > 0:
            keywords = keywords[:mk]

    platform = str(args.platform).strip() if args.platform else str(crawl_cfg.get("platform", "both")).strip() or "both"
    if platform not in {"boss", "zhilian", "both"}:
        raise SystemExit("platform 仅支持 boss/zhilian/both")

    if args.headless is None:
        headless = str_to_bool(str(crawl_cfg.get("headless", "true")))
    else:
        headless = str_to_bool(str(args.headless))
    verbose = str_to_bool(str(crawl_cfg.get("verbose", args.verbose)))
    user_data_dir = str(args.user_data_dir).strip() if args.user_data_dir else str(crawl_cfg.get("user_data_dir", "")).strip() or None
    if args.boss_headless is not None:
        boss_headless = str_to_bool(str(args.boss_headless))
    elif "boss_headless" in crawl_cfg:
        boss_headless = str_to_bool(str(crawl_cfg.get("boss_headless")))
    else:
        boss_headless = headless
    boss_channel = str(args.boss_channel).strip() if args.boss_channel else str(crawl_cfg.get("boss_channel", "")).strip() or None
    boss_cdp = str(args.boss_cdp).strip() if args.boss_cdp else str(crawl_cfg.get("boss_cdp", "")).strip() or None
    proxy = crawl_cfg.get("proxy", args.proxy)
    proxy = str(proxy).strip() if proxy else None
    slowmo = int(crawl_cfg.get("slowmo_ms", args.slowmo))
    sleep_min = float(crawl_cfg.get("sleep_min", args.sleep_min))
    sleep_max = float(crawl_cfg.get("sleep_max", args.sleep_max))
    out_path = str(args.out).strip() if args.out else str(crawl_cfg.get("output_excel", "")).strip() or None
    if not out_path:
        ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = os.path.join("output", f"salary_{ts}.xlsx")

    match_threshold = float(crawl_cfg.get("match_threshold", 0.4))
    match_relax_cfg = crawl_cfg.get("match_relax", None)
    llm_on = llm_enabled(llm_cfg)
    if match_relax_cfg is None:
        match_relax = bool(llm_on)
    else:
        match_relax = str_to_bool(str(match_relax_cfg))

    all_records: List[JobRecord] = []
    missing: List[Dict[str, Any]] = []
    if verbose:
        key_src = llm_get_api_key(llm_cfg)[1] if isinstance(llm_cfg, dict) else "none"
        llm_flag = "Y" if llm_on else "N"
        llm_model = str(llm_cfg.get("model", "")).strip() if isinstance(llm_cfg, dict) else ""
        print(f"[llm] cache_path={llm_cache_path} cached={len(_LLM_KEYWORD_PLAN_CACHE)}")
        print(f"[llm] enabled={llm_flag} model={llm_model or '-'} key_source={key_src}")
        print(
            f"[run] platform={platform} cities={len(cities)} pages_per_city={pages} keywords={len(keywords)} "
            f"headless={headless} boss_headless={boss_headless} proxy={'Y' if proxy else 'N'} "
            f"user_data_dir={user_data_dir or '-'} out={out_path}"
        )
    keyword_items: List[Dict[str, Any]] = []
    for kw in keywords:
        plan = llm_build_keyword_plan(kw, llm_cfg) if llm_on else {"search_query": "", "match_terms": []}
        search_kw = str(plan.get("search_query") or "").strip() or kw
        match_terms = plan.get("match_terms") if isinstance(plan.get("match_terms"), list) else None
        keyword_items.append({"keyword": kw, "search_keyword": search_kw, "match_terms": match_terms})
        if verbose and llm_on:
            print(f"[llm] kw={kw} search_kw={search_kw} match_terms={len(match_terms or [])}")
    if llm_on:
        llm_save_keyword_plan_cache(llm_cache_path, _LLM_KEYWORD_PLAN_CACHE)

    boss_gap_min = float(crawl_cfg.get("boss_gap_min", 12.0))
    boss_gap_max = float(crawl_cfg.get("boss_gap_max", 25.0))
    zhilian_gap_min = float(crawl_cfg.get("zhilian_gap_min", 1.5))
    zhilian_gap_max = float(crawl_cfg.get("zhilian_gap_max", 3.5))
    switch_gap_min = float(crawl_cfg.get("switch_gap_min", 6.0))
    switch_gap_max = float(crawl_cfg.get("switch_gap_max", 12.0))

    if boss_cdp and platform in {"boss", "both"}:
        boss_crawler = BossCrawler(
            headless=boss_headless,
            proxy=proxy,
            slow_mo_ms=slowmo,
            min_sleep=sleep_min,
            max_sleep=sleep_max,
            verbose=verbose,
            user_data_dir=user_data_dir,
            channel=boss_channel,
            cdp_endpoint=boss_cdp,
            random_pause=False,
            match_terms=None,
            match_threshold=match_threshold,
            match_relax=match_relax,
        )
        with sync_playwright() as p:
            if verbose:
                print(f"[boss] connect_cdp={boss_cdp}")
            boss_browser = p.chromium.connect_over_cdp(boss_cdp)
            boss_context = boss_browser.contexts[0] if boss_browser.contexts else boss_browser.new_context(**build_stealth_context_kwargs())
            ua = boss_get_user_agent_from_context(boss_context)
            api_page = boss_prepare_api_page(boss_context, verbose=verbose)

            zl_context = None
            zl_browser = None
            zl_page = None
            if platform in {"zhilian", "both"}:
                launch_kwargs: Dict[str, Any] = {"headless": headless, "slow_mo": slowmo}
                if proxy:
                    launch_kwargs["proxy"] = {"server": proxy}
                if user_data_dir:
                    zl_context = p.chromium.launch_persistent_context(user_data_dir, **launch_kwargs, **build_stealth_context_kwargs())
                    add_stealth_init_script(zl_context)
                else:
                    zl_browser = p.chromium.launch(**launch_kwargs)
                    zl_context = zl_browser.new_context(**build_stealth_context_kwargs())
                    add_stealth_init_script(zl_context)
                zl_page = zl_context.new_page()

            try:
                for item in keyword_items:
                    kw = str(item.get("keyword") or "").strip()
                    search_kw = str(item.get("search_keyword") or "").strip() or kw
                    match_terms = item.get("match_terms") if isinstance(item.get("match_terms"), list) else None
                    if not kw:
                        continue

                    boss_records, boss_missing = boss_crawler.crawl_cdp_one(
                        context=boss_context,
                        api_page=api_page,
                        ua=ua,
                        cities=cities,
                        keyword=kw,
                        search_keyword=search_kw,
                        match_terms=match_terms,
                        pages_per_city=pages,
                        boss_gap_min=boss_gap_min,
                        boss_gap_max=boss_gap_max,
                    )
                    all_records.extend(boss_records)
                    missing.extend(boss_missing)

                    if platform in {"zhilian", "both"} and zl_page is not None:
                        sleep_between(switch_gap_min, switch_gap_max, verbose=verbose, label="switch_boss_to_zhilian")
                        try:
                            zl_records = ZhilianCrawler(
                                headless=headless,
                                proxy=proxy,
                                slow_mo_ms=slowmo,
                                min_sleep=sleep_min,
                                max_sleep=sleep_max,
                                verbose=verbose,
                                user_data_dir=user_data_dir,
                                match_terms=match_terms,
                                match_threshold=match_threshold,
                                match_relax=match_relax,
                            ).crawl_with_page(
                                zl_page,
                                cities=cities,
                                keyword=kw,
                                pages_per_city=pages,
                                search_keyword=search_kw,
                                zhilian_gap_min=zhilian_gap_min,
                                zhilian_gap_max=zhilian_gap_max,
                            )
                            all_records.extend(zl_records)
                            city_hit = {r.city for r in zl_records}
                            for city in cities:
                                if city not in city_hit:
                                    missing.append({"platform": "zhilian", "city": city, "keyword": kw, "reason": "无结果"})
                        except Exception as e:
                            if verbose:
                                print(f"[zhilian] error kw={kw} err={e}")
                            for city in cities:
                                missing.append({"platform": "zhilian", "city": city, "keyword": kw, "reason": f"异常: {e}"})
                        sleep_between(switch_gap_min, switch_gap_max, verbose=verbose, label="switch_zhilian_to_boss")
            finally:
                if zl_context is not None:
                    try:
                        zl_context.close()
                    except Exception:
                        pass
                if zl_browser is not None:
                    try:
                        zl_browser.close()
                    except Exception:
                        pass
    else:
        for item in keyword_items:
            kw = str(item.get("keyword") or "").strip()
            search_kw = str(item.get("search_keyword") or "").strip() or kw
            match_terms = item.get("match_terms") if isinstance(item.get("match_terms"), list) else None

            if platform in {"boss", "both"}:
                try:
                    boss_records = BossCrawler(
                        headless=boss_headless,
                        proxy=proxy,
                        slow_mo_ms=slowmo,
                        min_sleep=sleep_min,
                        max_sleep=sleep_max,
                        verbose=verbose,
                        user_data_dir=user_data_dir,
                        channel=boss_channel,
                        cdp_endpoint=None,
                        random_pause=(args.keyword is not None and (not boss_headless)),
                        match_terms=match_terms,
                        match_threshold=match_threshold,
                        match_relax=match_relax,
                    ).crawl(cities=cities, keyword=kw, pages_per_city=pages, search_keyword=search_kw)
                    all_records.extend(boss_records)
                    city_hit = {r.city for r in boss_records}
                    for city in cities:
                        if city not in city_hit:
                            missing.append({"platform": "boss", "city": city, "keyword": kw, "reason": "无结果"})
                except Exception as e:
                    if verbose:
                        print(f"[boss] error kw={kw} err={e}")
                    for city in cities:
                        missing.append({"platform": "boss", "city": city, "keyword": kw, "reason": f"异常: {e}"})

            if platform in {"zhilian", "both"}:
                try:
                    zl_records = ZhilianCrawler(
                        headless=headless,
                        proxy=proxy,
                        slow_mo_ms=slowmo,
                        min_sleep=sleep_min,
                        max_sleep=sleep_max,
                        verbose=verbose,
                        user_data_dir=user_data_dir,
                        match_terms=match_terms,
                        match_threshold=match_threshold,
                        match_relax=match_relax,
                    ).crawl(cities=cities, keyword=kw, pages_per_city=pages, search_keyword=search_kw)
                    all_records.extend(zl_records)
                    city_hit = {r.city for r in zl_records}
                    for city in cities:
                        if city not in city_hit:
                            missing.append({"platform": "zhilian", "city": city, "keyword": kw, "reason": "无结果"})
                except Exception as e:
                    if verbose:
                        print(f"[zhilian] error kw={kw} err={e}")
                    for city in cities:
                        missing.append({"platform": "zhilian", "city": city, "keyword": kw, "reason": f"异常: {e}"})

    out_path = write_excel(all_records, out_path=out_path, missing=missing)
    if verbose:
        print(f"[run] excel_saved={out_path} records={len(all_records)} missing={len(missing)}")

    save_mysql = str_to_bool(str(mysql_cfg.get("enabled", args.save_mysql)))
    if save_mysql:
        table = str(mysql_cfg.get("table", args.mysql_table)).strip() or str(args.mysql_table)
        conn = mysql_connect_from_config(mysql_cfg) or mysql_connect_from_env()
        if not conn:
            raise SystemExit(
                "未检测到MySQL配置，请通过 config.yml 的 mysql 配置或设置 MYSQL_HOST/MYSQL_PORT/MYSQL_USER/MYSQL_PASSWORD/MYSQL_DATABASE"
            )
        try:
            mysql_ensure_table(conn, table=table)
            rows = mysql_insert_records(conn, table=table, records=all_records)
            if verbose:
                print(f"[mysql] table={table} rows_written={rows}")
        finally:
            conn.close()

    print(f"keywords={len(keywords)} records={len(all_records)} out={out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
